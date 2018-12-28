
from requsts_saver import saver
import trades_to_chart as ttc
import datetime as dt
import pandas as pd
import os
import re
import Refactoring_making_chart_func as mcf
from from_esignal import download_chart_of_ticker
import copy
import time

from plotly.offline import init_notebook_mode, plot
init_notebook_mode(connected=True)

class WrongDatesOrder(Exception):
    pass



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# In[33]:


def looking_for_chart(ticker, start_day, end_day=None, timeframe=1):
    if not end_day:
        end_day = start_day

    if start_day > end_day:
        raise WrongDatesOrder('Старт дей больше енд дея')

    try:
        files = os.listdir(os.path.join(esignal_charts_path, ticker))

    except FileNotFoundError:
        return None

    file_name_pattern = re.compile(r'[A-Z]+_(%s)_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2}).csv'%timeframe)
    existing_charts = {}

    for file in files:
        start_end = []
        parser = re.search(file_name_pattern, file)
        if not parser:
            continue

        start_end.extend([dt.datetime.strptime(parser[2], esignal_date_format),
                         dt.datetime.strptime(parser[3], esignal_date_format)])

        existing_charts[file] = start_end

    for file_name, [start_file, end_file] in existing_charts.items():

        if start_file <= start_day <= end_day <= end_file:
            return file_name

    return None


month_name_num_gen = {'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
                      'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12}

month_name_num_sbj = {'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5, 'июнь': 6, 'июль': 7,
                      'август': 8, 'сентябрь': 9, 'октябрь': 10, 'ноябрь': 11, 'декабрь': 12}

month_num_name_sbj = dict([(b, a.upper()) for a, b in month_name_num_sbj.items()])
month_num_name_gen = dict([(b, a.capitalize()) for a, b in month_name_num_gen.items()])

def preprocessing_best(storage_of_executions, folder_name):
    """парсинг Детейлд файла и вывод его в формат {дата : {тикер : датафррейм_екзекюшина}}"""
    file_parsing = {}
    excel_df = storage_of_executions
    reindex = {a: b for a, b in zip(range(22), excel_df.iloc[4, :])}  # вторая строка, 22 колонки в файле эксель
    excel_df = excel_df.rename(columns=reindex)

    date_re = re.compile(r'(\d{1,2})\s([а-яА-Я]+),\s(\d{4})')
    options_re = re.compile(r'^\+([\w/]+)\s{1,2}-\s.*')
    ticker_re = re.compile(r'^([\w/]+)\s{1,2}-\s.*')
    time_re = re.compile(r'\d\d:\d\d:\d\d')
    title_re = re.compile(r'Результат № (\w{1,2})')

    options = False
    for i in range(len(excel_df)):
        if pd.notna(excel_df.iloc[i][0]):
            if options == False:
                if re.match(title_re, excel_df.iloc[i][0]):
                    result_num = int(re.match(title_re, excel_df.iloc[i][0])[1])

                elif re.match(date_re, excel_df.iloc[i][0]):  # дата
                    search = re.search(date_re, excel_df.iloc[i][0])
                    date = f'{month_name_num_gen[search[2].lower()]}/{int(search[1])}/{search[3]}'

                    if date not in file_parsing.keys():
                        file_parsing[date] = {}

                    continue

                #elif re.match(options_re, excel_df.iloc[i][0]):  # опционы
                #    options = True
                #    continue

                #elif re.match(ticker_re, excel_df.iloc[i][0]):  # если тикер
                #    ticker = re.match(ticker_re, excel_df.iloc[i][0])[1]
                #    #file_parsing[date][ticker] = pd.DataFrame()
                #    continue

                elif excel_df.iloc[i][0] == 'Time':
                    if re.match(options_re, excel_df.iloc[i-1][0]):  # опционы
                        options = True
                        continue

                    elif re.match(ticker_re, excel_df.iloc[i-1][0]):  # если тикер
                        ticker = re.match(ticker_re, excel_df.iloc[i-1][0])[1]

                    else:
                        ticker = 'NonFormatTicker'

                    file_parsing[date][ticker] = pd.DataFrame()
                    file_parsing[date][ticker].reindex(columns=excel_df.iloc[i])

                elif re.match(time_re, excel_df.iloc[i][0]):  # Время
                    file_parsing[date][ticker] = file_parsing[date][ticker].append(excel_df.iloc[i, :])

            else:
                if re.match(date_re, excel_df.iloc[i][0]):
                    options = False
                    date = excel_df.iloc[i][0]
                    file_parsing[date] = {}

                elif re.match(ticker_re, excel_df.iloc[i][0]):
                    options = False
                    ticker = re.match(ticker_re, excel_df.iloc[i][0])[1]
                    file_parsing[date][ticker] = pd.DataFrame()

    date_ticker_exec = ttc.normalise_formats(file_parsing)
    return date_ticker_exec


def make_correct_format_dict(path, date, drop_wrong_ticker=True):
    folder_date_month = f'{str(date.month).zfill(2)} {month_num_name_sbj[date.month]} {date.year}'
    folder_date_day = f'{str(date.day).zfill(2)} {month_num_name_gen[date.month]} {date.year}'
    main_dict = {}
    for folder in os.listdir(path):
        try:
            for file in os.listdir(os.path.join(path, folder, folder_date_month, folder_date_day)):
                storage_of_exec = ttc.open_excel(os.path.join(path, folder, folder_date_month, folder_date_day, file))
                preproc_dictionary = preprocessing_best(storage_of_exec, folder)  # возвращает Дата : Тикер : Датафрейм
                for date in preproc_dictionary.keys():
                    for ticker in preproc_dictionary[date]:
                        main_dict.setdefault(date, {}).setdefault(ticker, {}).update(
                            {file.split('- ')[-1][:-4]: preproc_dictionary[date][ticker]})
        except FileNotFoundError:
            pass
    if drop_wrong_ticker:
        for date in main_dict.keys():
            main_dict[date].pop('NonFormatTicker', None)

    return main_dict


def trades_more_amount(date_ticker_account_exec, amount=500):
    date_ticker_acc_exec = copy.deepcopy(date_ticker_account_exec)
    for date in date_ticker_account_exec.keys():
        for ticker in date_ticker_account_exec[date].keys():
            for acc in date_ticker_account_exec[date][ticker].keys():
                try:
                    if date_ticker_account_exec[date][ticker][acc].empty or\
                            date_ticker_account_exec[date][ticker][acc].Gross.sum() < amount:
                        date_ticker_acc_exec[date][ticker].pop(acc)
                        if not date_ticker_acc_exec[date][ticker]:
                            date_ticker_acc_exec[date].pop(ticker)
                            if not date_ticker_acc_exec[date]:
                                date_ticker_acc_exec.pop(date)
                except AttributeError:
                    print(date, ticker, acc)
                    raise AttributeError

    return date_ticker_acc_exec


propreports_file_date_format = '%Y-%m-%d'
esignal_date_format = '%Y-%m-%d'
esignal_charts_path = r'E:\Trading_diary\Esignal_charts'
bests_path = r'E:\Trading_diary\Folder'
detailed_folder = r'E:\Trading_diary\Detailed'
login = '07060017'
password = 'elisabet777'


def main():
    startDate = '2018-12-18'
    dt_startDate = dt.datetime.strptime(startDate, propreports_file_date_format)
    endDate = '2018-12-18'
    dt_endDate = dt.datetime.strptime(endDate, propreports_file_date_format)


    # парсим файлы статистики
    '''main_dict = {}
    for folder in os.listdir(bests_folder):
        for file in os.listdir(folder):
            path = os.path.join(bests_folders, folder, file)
            opened_file = ttc.open_excel(path)

            main_dict = {**main_dict, **preprocessing_best(opened_file, folder)}'''

    main_dict = {}
    while dt_startDate <= dt_endDate:
        print(dt.datetime.now(), f' : creating main dict {dt_startDate}')
        main_dict = {**main_dict, **trades_more_amount(make_correct_format_dict(bests_path, dt_startDate), amount=500)}
        dt_startDate += dt.timedelta(days=1)

    # сделать цикл по всему дикту
    for date in main_dict:
        dt_date = dt.datetime.strptime(date, '%m/%d/%Y')
        spy_file_name = looking_for_chart('SPY', dt_date)

        if not spy_file_name:
            spy_daily_file = download_chart_of_ticker('SPY', dt_date, 'd')
            spy_file_name = download_chart_of_ticker('SPY', dt_date, 1)

        else:
            spy_daily_file = 'SPY_latest_daily.csv'

        spy_intraday_df = ttc.create_chart_df(os.path.join(esignal_charts_path, 'SPY', spy_file_name))
        needed_spy_chart = ttc.chart_with_needed_dates(spy_intraday_df, dt_date)
        spy_daily_df = ttc.create_chart_df(os.path.join(esignal_charts_path, 'SPY', spy_daily_file))
        needed_spy_daily_chart = spy_daily_df.loc[spy_daily_df.Date_Time <= dt_date]

        for ticker in main_dict[date]:
            print(dt.datetime.now(), f'work with date_ticker_acc_exec, date: {date}, ticker : {ticker}')
            ticker_file_name = looking_for_chart(ticker, dt_date, timeframe=1)

            # проверить, есть ли данный график в скачанных экселях, если нет, то скачать
            if not ticker_file_name:
                ticker_daily_file = download_chart_of_ticker(ticker, dt_date, 'd')

                time.sleep(1)
                if os.path.split(ticker_daily_file)[-1] not in os.listdir(os.path.join(esignal_charts_path, ticker)):
                    continue

                ticker_file_name = download_chart_of_ticker(ticker, dt_date, 1)



            else:
                ticker_daily_file = f'{ticker}_latest_daily.csv'

            stock_df = ttc.create_chart_df(os.path.join(esignal_charts_path, ticker, ticker_file_name))
            needed_stock_chart = ttc.chart_with_needed_dates(stock_df, dt_date)
            if needed_stock_chart.empty:
                print(date, ticker, acc)

            stock_daily_df = ttc.create_chart_df(os.path.join(esignal_charts_path, ticker, ticker_daily_file))
            needed_stock_daily_chart = stock_daily_df.loc[stock_daily_df.Date_Time <= dt_date]


            # создаем трейсы графика

            chart_traces = mcf.preparation_chart_traces(needed_stock_chart, needed_stock_daily_chart,
                                                        needed_spy_chart, needed_spy_daily_chart, ticker)

            # создаем трейсы трейдов
            trades_traces = []
            for acc, symbol in zip(main_dict[date][ticker], mcf.symbols):
                trades_traces.extend(mcf.preparation_trades_traces(acc, main_dict[date][ticker][acc],
                                                                   symbol))

            # лаяут
            layout = mcf.make_layout(dt_date, needed_stock_chart, exec_dict=main_dict[date][ticker])

            data = [*chart_traces, *trades_traces]

            net = main_dict[date][ticker][mcf.find_max_net_key(main_dict[date][ticker])].Net.sum()
            filename = r'E:\Trading_diary\Drawn_charts\Bests' + f'\{dt_date.strftime("%Y-%m-%d")}_{ticker}_{int(net)}.html'

            plot(dict(data=data, layout=layout), auto_open=False, filename=filename)

            for acc, df in main_dict[date][ticker].items():
                line = pd.DataFrame([{'Date': date, 'Result_Num': acc, 'Stock_Opt': 'stock', 'Ticker': ticker,
                                      'Start_time': df.Date_Time.iloc[0].time(),
                                      'End_time': df.Date_Time.iloc[-1].time(),
                                      'Number_of_trades' : len(df.loc[df.Pos_Size == 0]),
                                      'Gross': df.Gross.sum(),
                                      'Vol': df.Qty.abs().sum(),
                                      'Net': df.Net.sum(),
                                      'Chart': filename,
                                      'Comment': ''}],
                                        columns=['Date', 'Result_Num', 'Stock_Opt', 'Ticker', 'Start_time', 'End_time',
                                                 'Number_of_trades', 'Gross', 'Vol', 'Net', 'Chart', 'Comment'])

                if 'Diary.xlsx' in os.listdir(r'E:\Trading_diary\Drawn_charts\Bests'):
                    append_df_to_excel(r'E:\Trading_diary\Drawn_charts\Bests\Diary.xlsx', line, header=None, index=False)

                else:
                    append_df_to_excel(r'E:\Trading_diary\Drawn_charts\Bests\Diary.xlsx', line, index=False)

    return main_dict


if __name__ == '__main__':
    dictionary = main()

