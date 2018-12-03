
# coding: utf-8

# Скачать статистику
# распарсить ее
# определить, есть ли график данного тикера и спайдера. если нет, то скачать
# создать графики и дневник
# (https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas)
#

# In[28]:


from requsts_saver import saver
import trades_to_chart as ttc
import datetime as dt
import pandas as pd
import os
import re
#from downloader_ticker_charts import download_chart_of_ticker
from from_esignal import download_chart_of_ticker
from collections import OrderedDict

import pdb

# In[31]:


class WrongDatesOrder(Exception):
    pass


# In[32]:


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# In[33]:


def looking_for_chart(ticker, start_day, end_day=None, timeframe=1):
    if not end_day:
        end_day = start_day

    if start_day > end_day:
        raise WrongDatesOrder('Старт дей больше енд дея')

    try:
        files = os.listdir(os.path.join(esignal_charts_path, ticker))

    except FileNotFoundError:
        return None

    file_name_pattern = re.compile(r'[A-Z]+_(%s)_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2}).csv'%timeframe)
    existing_charts = {}

    for file in files:
        start_end = []
        parser = re.search(file_name_pattern, file)
        if not parser:
            continue

        start_end.extend([dt.datetime.strptime(parser[2], esignal_date_format),
                         dt.datetime.strptime(parser[3], esignal_date_format)])

        existing_charts[file] = start_end

    for file_name, [start_file, end_file] in existing_charts.items():

        if start_file <= start_day <= end_day <= end_file:
            return file_name

    return None



# In[35]:


propreports_file_date_format = '%Y-%m-%d'
esignal_date_format = '%Y-%m-%d'
esignal_charts_path = r'E:\Trading_diary\Esignal_charts'
detailed_folder = r'E:\Trading_diary\Detailed'
login = '07060212'
password = 'parolo12'

def main():
    startDate = '2018-07-05'
    dt_startDate = dt.datetime.strptime(startDate, propreports_file_date_format)
    endDate = '2018-07-06'
    dt_endDate = dt.datetime.strptime(endDate, propreports_file_date_format)

    # скачиваем файлы статистики
    downloaded_detailed = saver(login, password, dt_startDate, dt_endDate)

    # парсим файлы статистики
    main_dict = {}
    for file in downloaded_detailed:
        path = os.path.join(detailed_folder, file)
        opened_file = ttc.open_excel(path)

        main_dict = {**main_dict, **ttc.preprocessing_trades(opened_file)}

    # сделать цикл по всему дикту
    for date in main_dict:
        dt_date = dt.datetime.strptime(date, '%m/%d/%Y')
        spy_file_name = looking_for_chart('SPY', dt_date)

        if not spy_file_name:
            spy_daily_file = download_chart_of_ticker('SPY', dt_date, 'd')
            spy_file_name = download_chart_of_ticker('SPY', dt_date, 1)

        else:
            spy_daily_file = 'SPY_latest_daily.csv'

        spy_intraday_df = ttc.create_chart_df(os.path.join(esignal_charts_path, 'SPY', spy_file_name))
        needed_spy_chart = ttc.chart_with_needed_dates(spy_intraday_df, dt_date)
        spy_daily_df = ttc.create_chart_df(os.path.join(esignal_charts_path, 'SPY', spy_daily_file))
        needed_spy_daily_chart = spy_daily_df.loc[spy_daily_df.Date_Time <= dt_date]

        for ticker in main_dict[date]:
            print(dt_date)
            ticker_file_name = looking_for_chart(ticker, dt_date, timeframe=1)

            # проверить, есть ли данный график в скачанных экселях, если нет, то скачать
            if not ticker_file_name:
                ticker_daily_file = download_chart_of_ticker(ticker, dt_date, 'd')
                ticker_file_name = download_chart_of_ticker(ticker, dt_date, 1)

            else:
                ticker_daily_file = f'{ticker}_latest_daily.csv'
                #pdb.set_trace()

            stock_df = ttc.create_chart_df(os.path.join(esignal_charts_path, ticker, ticker_file_name))
            needed_stock_chart = ttc.chart_with_needed_dates(stock_df, dt_date)

            stock_daily_df = ttc.create_chart_df(os.path.join(esignal_charts_path, ticker, ticker_daily_file))
            needed_stock_daily_chart = stock_daily_df.loc[stock_daily_df.Date_Time <= dt_date]
            execution = main_dict[date][ticker]
            path_to_created_chart = ttc.make_main_chart(needed_chart=needed_stock_chart,
                                                        daily_df=needed_stock_daily_chart,
                                                        drawing_trades=execution,
                                                        spy_needed_chart=needed_spy_chart,
                                                        spy_daily=needed_spy_daily_chart)

            line = pd.DataFrame([{'Date': date, 'Stock_Opt': 'stock', 'Ticker': ticker,
                              'Start_time': execution.Date_Time.iloc[0].time(),
                              'End_time' : execution.Date_Time.iloc[-1].time(),
                              'Number_of_trades' : len(execution.loc[execution.Pos_Size == 0]),
                              'Gross' : execution.Gross.sum(),
                              'Vol' : execution.Qty.abs().sum(),
                              'Net' : execution.Net.sum(),
                              'Chart' : path_to_created_chart,
                              'Comment' : ''}],
                                columns=['Date', 'Stock_Opt', 'Ticker', 'Start_time', 'End_time', 'Number_of_trades', 'Gross', 'Vol', 'Net', 'Chart', 'Comment'])

            if 'Diary.xlsx' in os.listdir(r'E:\Trading_diary'):
                append_df_to_excel(r'E:\Trading_diary\Diary.xlsx', line, header=None, index=False)

            else:
                append_df_to_excel(r'E:\Trading_diary\Diary.xlsx', line, index=False)



# In[36]:


main()

